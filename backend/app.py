"""婚礼宾客统计系统 — FastAPI 后端。

数据存 data/wedding.db（SQLite）；宾客名单支持 xlsx 导入/导出。
"""

import re
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

import storage
from storage import CONFIRM_STATUSES, INVITE_STATUSES, WECHAT_STATUSES

GUEST_TYPES = ["宾客及家属", "宾客", "家属"]

app = FastAPI(title="婚礼宾客统计系统")

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"

storage.ensure_data_files()


# ---------- 请求模型 ----------

class GuestIn(BaseModel):
    name: str = Field(min_length=1)
    party_size: int = Field(default=1, ge=1)          # 预计人数（含本人）
    confirmed_size: int | None = Field(default=None, ge=0)  # 确认会到人数，None = 等于预计人数
    family_names: str = ""
    table_no: str = ""
    invite_status: str = "未发送"
    confirm_status: str = "待确认"
    wechat_sent: str = "未发送"
    note: str = ""
    guest_type: str = "宾客及家属"
    force: bool = False  # 目标桌超容时是否强制安排


class MoveIn(BaseModel):
    target_table: str = ""  # 空字符串 = 移到未安排区
    with_family: bool = True
    force: bool = False


class TableIn(BaseModel):
    table_no: str = Field(min_length=1)
    label: str = ""
    capacity: int | None = Field(default=None, ge=1)  # None = 使用全局默认容量
    x: float | None = Field(default=None, ge=0)       # 桌心坐标（米），None = 未摆放
    y: float | None = Field(default=None, ge=0)


class BatchUpdateIn(BaseModel):
    ids: list[int] = Field(min_length=1)
    table_no: str | None = None
    invite_status: str | None = None
    confirm_status: str | None = None
    note: str | None = None
    wechat_sent: str | None = None


class ConfigIn(BaseModel):
    default_capacity: int = Field(ge=1)
    budget_total: int = Field(ge=0)
    venue_width: float = Field(default=18.0, gt=0, le=200)   # 会场宽（米，横向）
    venue_depth: float = Field(default=25.0, gt=0, le=200)   # 会场长（米，纵向）
    table_diameter: float = Field(default=1.8, gt=0, le=10)  # 桌子直径（米）
    table_gap: float = Field(default=1.2, ge=0, le=20)       # 自动排列桌边间距（米）
    wedding_date: str = Field(default="10.25", max_length=20)


# ---------- 内部工具 ----------

def _validate_statuses(invite_status: str, confirm_status: str, wechat_sent: str):
    if invite_status not in INVITE_STATUSES:
        raise HTTPException(400, f"邀请函状态必须是 {INVITE_STATUSES} 之一")
    if confirm_status not in CONFIRM_STATUSES:
        raise HTTPException(400, f"确认状态必须是 {CONFIRM_STATUSES} 之一")
    if wechat_sent not in WECHAT_STATUSES:
        raise HTTPException(400, f"微信通知状态必须是 {WECHAT_STATUSES} 之一")


def _seat_size(g: dict) -> int:
    """占座人数：已确认按确认人数，待确认按预计人数，不参加不占座。"""
    if g["confirm_status"] == "不参加":
        return 0
    if g["confirm_status"] == "已确认":
        return g["confirmed_size"]
    return g["party_size"]


def _table_occupied(guests: list[dict], table_no: str, exclude_ids: set[int] = frozenset()) -> int:
    return sum(
        _seat_size(g) for g in guests
        if g["table_no"] == table_no and g["id"] not in exclude_ids
    )


def _effective_capacity(table: dict, config: dict) -> int:
    return table["capacity"] if table["capacity"] is not None else config["default_capacity"]


def _check_capacity(guests, tables, config, table_no: str, incoming: int,
                    exclude_ids: set[int] = frozenset(), force: bool = False):
    """校验目标桌容量。桌号不存在报 404，超容且未 force 报 409。"""
    if not table_no:
        return
    table = next((t for t in tables if t["table_no"] == table_no), None)
    if table is None:
        raise HTTPException(404, f"桌号「{table_no}」不存在")
    if force:
        return
    cap = _effective_capacity(table, config)
    occupied = _table_occupied(guests, table_no, exclude_ids)
    if occupied + incoming > cap:
        raise HTTPException(
            409,
            f"桌「{table_no}」容量不足：容纳 {cap} 人，已坐 {occupied} 人，"
            f"再安排 {incoming} 人将超出。可强制安排。",
        )


def _find_guest(guests: list[dict], gid: int) -> dict:
    guest = next((g for g in guests if g["id"] == gid), None)
    if guest is None:
        raise HTTPException(404, f"宾客 ID {gid} 不存在")
    return guest


def _build_stats(guests: list[dict], config: dict) -> dict:
    # 统计栏展示录入的预计与确认总量；占座人数仍由 _seat_size 按状态计算。
    expected = sum(g["party_size"] for g in guests)
    confirmed = sum(g["confirmed_size"] for g in guests)
    pending = sum(g["party_size"] for g in guests if g["confirm_status"] == "待确认")
    declined = sum(g["party_size"] for g in guests if g["confirm_status"] == "不参加")
    return {
        "budget_total": config["budget_total"],
        "invited_total": sum(g["party_size"] for g in guests),  # 预算合计
        "confirmed": confirmed,
        "pending": pending,
        "declined": declined,
        "expected": expected,
        "seated": sum(_seat_size(g) for g in guests if g["table_no"]),
        "unseated": sum(_seat_size(g) for g in guests if not g["table_no"]),
        "guest_count": len(guests),
        "invite_sent": sum(1 for g in guests if g["invite_status"] == "已发送"),
        "invite_unsent": sum(1 for g in guests if g["invite_status"] == "未发送"),
        "wechat_sent": sum(1 for g in guests if g.get("wechat_sent") == "已发送"),
    }


# ---------- 总览 ----------

@app.get("/api/overview")
def overview():
    guests = storage.load_guests()
    tables = storage.load_tables()
    config = storage.load_config()
    table_views = []
    for t in tables:
        members = [g for g in guests if g["table_no"] == t["table_no"]]
        table_views.append({
            **t,
            "effective_capacity": _effective_capacity(t, config),
            "occupied": _table_occupied(guests, t["table_no"]),
            "guests": members,
        })
    known_tables = {t["table_no"] for t in tables}
    unassigned = [g for g in guests if not g["table_no"] or g["table_no"] not in known_tables]
    return {
        "guests": guests,
        "tables": table_views,
        "unassigned_guests": unassigned,
        "config": config,
        "stats": _build_stats(guests, config),
    }


# ---------- 宾客 ----------

def _validate_guest_type(guest_type: str, party_size: int):
    if guest_type not in GUEST_TYPES:
        raise HTTPException(400, f"宾客类型必须是 {GUEST_TYPES} 之一")
    if guest_type == "宾客" and party_size != 1:
        raise HTTPException(400, "「宾客」类型的人数固定为 1")


def _next_group_id(guests: list[dict]) -> int:
    return max((g.get("group_id") or 0 for g in guests), default=0) + 1


def _guest_fields(body: GuestIn) -> dict:
    return {
        "name": body.name.strip(),
        "party_size": body.party_size,
        "confirmed_size": body.confirmed_size if body.confirmed_size is not None else body.party_size,
        "family_names": body.family_names.strip(),
        "table_no": body.table_no.strip(),
        "invite_status": body.invite_status,
        "confirm_status": body.confirm_status,
        "wechat_sent": body.wechat_sent,
        "note": body.note.strip(),
        "guest_type": body.guest_type,
    }


@app.post("/api/guests")
def create_guest(body: GuestIn):
    _validate_statuses(body.invite_status, body.confirm_status, body.wechat_sent)
    _validate_guest_type(body.guest_type, body.party_size)
    guests = storage.load_guests()
    tables = storage.load_tables()
    config = storage.load_config()
    fields = _guest_fields(body)
    guest = {"id": storage.next_guest_id(guests), "group_id": None, **fields}
    _check_capacity(guests, tables, config, guest["table_no"], _seat_size(guest), force=body.force)
    guests.append(guest)
    storage.save_guests(guests)
    return guest


@app.put("/api/guests/{gid}")
def update_guest(gid: int, body: GuestIn):
    _validate_statuses(body.invite_status, body.confirm_status, body.wechat_sent)
    _validate_guest_type(body.guest_type, body.party_size)
    guests = storage.load_guests()
    tables = storage.load_tables()
    config = storage.load_config()
    guest = _find_guest(guests, gid)
    fields = _guest_fields(body)
    _check_capacity(guests, tables, config, fields["table_no"], _seat_size(fields),
                    exclude_ids={gid}, force=body.force)
    guest.update(fields)
    storage.save_guests(guests)
    return guest


@app.delete("/api/guests/{gid}")
def delete_guest(gid: int):
    guests = storage.load_guests()
    guest = _find_guest(guests, gid)
    guests.remove(guest)
    storage.save_guests(guests)
    return {"ok": True}


@app.patch("/api/guests/batch")
def batch_update_guests(body: BatchUpdateIn):
    if body.invite_status is not None and body.invite_status not in INVITE_STATUSES:
        raise HTTPException(400, f"invite_status 必须是 {'、'.join(INVITE_STATUSES)} 之一")
    if body.confirm_status is not None and body.confirm_status not in CONFIRM_STATUSES:
        raise HTTPException(400, f"confirm_status 必须是 {'、'.join(CONFIRM_STATUSES)} 之一")
    if body.wechat_sent is not None and body.wechat_sent not in WECHAT_STATUSES:
        raise HTTPException(400, f"wechat_sent 必须是 {'、'.join(WECHAT_STATUSES)} 之一")
    guests = storage.load_guests()
    id_set = set(body.ids)
    updated = 0
    for g in guests:
        if g["id"] not in id_set:
            continue
        if body.table_no is not None:
            g["table_no"] = body.table_no
        if body.invite_status is not None:
            g["invite_status"] = body.invite_status
        if body.confirm_status is not None:
            g["confirm_status"] = body.confirm_status
        if body.note is not None:
            g["note"] = body.note
        if body.wechat_sent is not None:
            g["wechat_sent"] = body.wechat_sent
        updated += 1
    storage.save_guests(guests)
    return {"updated": updated}


@app.post("/api/guests/{gid}/move")
def move_guest(gid: int, body: MoveIn):
    """换桌。with_family=False 且「宾客及家属」类型人数 > 1 时拆分记录。"""
    guests = storage.load_guests()
    tables = storage.load_tables()
    config = storage.load_config()
    guest = _find_guest(guests, gid)
    target = body.target_table.strip()

    can_split = (guest.get("guest_type", "宾客及家属") == "宾客及家属"
                 and guest["party_size"] > 1)
    split = not body.with_family and can_split
    if split:
        person_confirmed = min(1, guest["confirmed_size"])
        moving = {**guest, "party_size": 1, "confirmed_size": person_confirmed}
    else:
        moving = guest
    _check_capacity(guests, tables, config, target, _seat_size(moving),
                    exclude_ids={gid}, force=body.force)

    result = {"moved": None, "family_left": None}
    if split:
        group_id = _next_group_id(guests)
        family = {
            "id": storage.next_guest_id(guests),
            "name": f"{guest['name']}家属",
            "party_size": guest["party_size"] - 1,
            "confirmed_size": guest["confirmed_size"] - person_confirmed,
            "family_names": guest["family_names"],
            "table_no": guest["table_no"],
            "invite_status": guest["invite_status"],
            "wechat_sent": guest.get("wechat_sent", "未发送"),
            "confirm_status": guest["confirm_status"],
            "note": "",
            "guest_type": "家属",
            "group_id": group_id,
        }
        guests.append(family)
        guest.update({
            "party_size": 1, "confirmed_size": person_confirmed,
            "family_names": "", "table_no": target,
            "guest_type": "宾客", "group_id": group_id,
        })
        result["family_left"] = family
    else:
        guest["table_no"] = target
    result["moved"] = guest
    storage.save_guests(guests)
    return result


@app.post("/api/guests/{gid}/split")
def split_guest(gid: int):
    """将「宾客及家属」拆分为「宾客」(1人) + 「家属」(N-1人)，共享 group_id。"""
    guests = storage.load_guests()
    guest = _find_guest(guests, gid)
    if guest.get("guest_type", "宾客及家属") != "宾客及家属":
        raise HTTPException(400, "只有「宾客及家属」类型可以拆分")
    if guest["party_size"] <= 1:
        raise HTTPException(400, "人数为 1 的宾客无法拆分")

    group_id = _next_group_id(guests)
    person_confirmed = min(1, guest["confirmed_size"])
    family = {
        "id": storage.next_guest_id(guests),
        "name": f"{guest['name']}家属",
        "party_size": guest["party_size"] - 1,
        "confirmed_size": guest["confirmed_size"] - person_confirmed,
        "family_names": guest["family_names"],
        "table_no": guest["table_no"],
        "invite_status": guest["invite_status"],
        "wechat_sent": guest.get("wechat_sent", "未发送"),
        "confirm_status": guest["confirm_status"],
        "note": guest["note"],
        "guest_type": "家属",
        "group_id": group_id,
    }
    guests.append(family)
    guest.update({
        "party_size": 1, "confirmed_size": person_confirmed,
        "family_names": "", "guest_type": "宾客", "group_id": group_id,
    })
    storage.save_guests(guests)
    return {"guest": guest, "family": family}


@app.post("/api/guests/{gid}/merge")
def merge_guest(gid: int):
    """将同 group_id 的「宾客」+「家属」合并回一条「宾客及家属」。"""
    guests = storage.load_guests()
    guest = _find_guest(guests, gid)
    gid_group = guest.get("group_id")
    if not gid_group:
        raise HTTPException(400, "该宾客没有关联的拆分组")
    group = [g for g in guests if g.get("group_id") == gid_group]
    if len(group) != 2:
        raise HTTPException(400, "拆分组数据异常，无法合并")
    primary = next((g for g in group if g.get("guest_type") == "宾客"), None)
    family = next((g for g in group if g.get("guest_type") == "家属"), None)
    if not primary or not family:
        raise HTTPException(400, "拆分组缺少宾客或家属记录")
    primary.update({
        "party_size": primary["party_size"] + family["party_size"],
        "confirmed_size": primary["confirmed_size"] + family["confirmed_size"],
        "family_names": family["family_names"],
        "guest_type": "宾客及家属",
        "group_id": None,
    })
    guests.remove(family)
    storage.save_guests(guests)
    return {"merged": primary}


# ---------- 宾客名单导入 / 导出（xlsx） ----------
# 交换格式：姓名 | 家属 | 预计人数 | 确认人数 | 桌号 | 邀请状态 | 确认状态 | 备注

EXPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _norm_cell(v) -> str:
    """单元格转字符串；Excel 数字桌号 1.0 规整为 '1'。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _cell_int(v) -> int | None:
    s = _norm_cell(v)
    try:
        return int(float(s)) if s else None
    except ValueError:
        return None


_RE_NAME_ETC = re.compile(r"^(.+?)等(\d+)位家属$")
_RE_COUNT_ONLY = re.compile(r"^(\d+)位家属$")


def _format_family(names: list[str], party_size: int) -> str:
    family_count = party_size - 1
    if family_count <= 0:
        return ""
    if not names:
        return f"{family_count}位家属"
    if len(names) >= family_count:
        return "、".join(names[:family_count])
    return f"{'、'.join(names)}等{family_count}位家属"


def _parse_family_names(text: str) -> list[str]:
    text = text.strip()
    if not text or text == "无":
        return []
    if _RE_COUNT_ONLY.match(text):
        return []
    m = _RE_NAME_ETC.match(text)
    if m:
        return [n.strip() for n in m.group(1).split("、") if n.strip()]
    return [n.strip() for n in text.split("、") if n.strip()]


def _export_family(g) -> str:
    fn = g["family_names"]
    if not fn:
        return "无"
    if "," in fn:
        names = [s.strip() for s in fn.split(",") if s.strip()]
    else:
        names = _parse_family_names(fn)
    return _format_family(names, g["party_size"]) or "无"


@app.get("/api/guests/export")
def export_guests():
    guests = storage.load_guests()
    headers = ["姓名", "家属", "预计人数", "确认人数", "桌号",
               "邀请状态", "微信通知", "确认状态", "备注", "类型"]

    wb = Workbook()
    ws = wb.active
    ws.title = "宾客名单"
    ws.append(headers)
    for g in guests:
        ws.append([g["name"], _export_family(g), g["party_size"], g["confirmed_size"],
                   g["table_no"], g["invite_status"], g.get("wechat_sent", "未发送"),
                   g["confirm_status"], g["note"], g.get("guest_type", "宾客及家属")])

    # 表头样式
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="A2687B", end_color="A2687B", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    # 数据行样式
    row_border = Border(bottom=Side(style="thin", color="D3BFB0"))
    row_align = Alignment(vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = row_align
            cell.border = row_border

    # 列宽自适应
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 8), 40)

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(), media_type=EXPORT_MIME, headers={
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote('宾客名单.xlsx')}",
    })


@app.post("/api/guests/import")
async def import_guests(request: Request, mode: str = "append"):
    """导入宾客名单。body 为 xlsx 原始字节；mode=append 追加 / replace 覆盖全部。"""
    if mode not in ("append", "replace"):
        raise HTTPException(400, "mode 必须是 append 或 replace")
    data = await request.body()
    if not data:
        raise HTTPException(400, "请上传 xlsx 文件")
    try:
        ws = load_workbook(BytesIO(data), data_only=True).active
    except Exception:
        raise HTTPException(400, "无法解析文件，请上传 .xlsx 格式的 Excel 文件")

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or ()
    col = {}
    for i, h in enumerate(header):
        hname = str(h or "").strip()
        if hname:
            col[hname] = i
    if "姓名" not in col:
        raise HTTPException(400, "缺少「姓名」列（表头需含：姓名、家属、预计人数、确认人数、桌号）")

    def cell(vals, cname):
        i = col.get(cname)
        return vals[i] if i is not None and i < len(vals) else None

    guests = [] if mode == "replace" else storage.load_guests()
    next_id = storage.next_guest_id(guests)
    tables = storage.load_tables()
    known_tables = {t["table_no"] for t in tables}
    imported, skipped, unknown_tables = 0, 0, []

    for row in rows:
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = _norm_cell(cell(row, "姓名"))
        if not name:
            skipped += 1
            continue

        party = _cell_int(cell(row, "预计人数")) or 1

        family_text = _norm_cell(cell(row, "家属"))
        family_names_list = _parse_family_names(family_text)
        family_names = _format_family(family_names_list, party)

        conf_raw = _cell_int(cell(row, "确认人数"))
        confirmed = min(conf_raw, party) if conf_raw is not None else party

        table_no = _norm_cell(cell(row, "桌号"))
        if table_no and table_no not in known_tables:
            unknown_tables.append(table_no)
            known_tables.add(table_no)

        invite_raw = _norm_cell(cell(row, "邀请状态"))
        invite_status = invite_raw if invite_raw in INVITE_STATUSES else "未发送"

        confirm_raw = _norm_cell(cell(row, "确认状态"))
        confirm_status = confirm_raw if confirm_raw in CONFIRM_STATUSES else "待确认"

        wechat_raw = _norm_cell(cell(row, "微信通知"))
        wechat_sent = wechat_raw if wechat_raw in WECHAT_STATUSES else "未发送"

        note = _norm_cell(cell(row, "备注"))

        type_raw = _norm_cell(cell(row, "类型"))
        guest_type = type_raw if type_raw in GUEST_TYPES else "宾客及家属"

        guests.append({
            "id": next_id,
            "name": name,
            "party_size": party,
            "confirmed_size": max(0, confirmed),
            "family_names": family_names,
            "table_no": table_no,
            "invite_status": invite_status,
            "wechat_sent": wechat_sent,
            "confirm_status": confirm_status,
            "note": note,
            "guest_type": guest_type,
            "group_id": None,
        })
        next_id += 1
        imported += 1

    if imported == 0 and skipped == 0:
        raise HTTPException(400, "文件里没有可导入的数据行")
    for no in unknown_tables:
        tables.append({"table_no": no, "label": "", "capacity": None, "x": None, "y": None})
    if unknown_tables:
        storage.save_tables(tables)
    storage.save_guests(guests)
    return {"imported": imported, "skipped": skipped,
            "created_tables": unknown_tables, "mode": mode}


# ---------- 桌子 ----------

@app.post("/api/tables")
def create_table(body: TableIn):
    tables = storage.load_tables()
    no = body.table_no.strip()
    if any(t["table_no"] == no for t in tables):
        raise HTTPException(409, f"桌号「{no}」已存在")
    table = {"table_no": no, "label": body.label.strip(), "capacity": body.capacity,
             "x": body.x, "y": body.y}
    tables.append(table)
    storage.save_tables(tables)
    return table


@app.put("/api/tables/{table_no}")
def update_table(table_no: str, body: TableIn):
    tables = storage.load_tables()
    table = next((t for t in tables if t["table_no"] == table_no), None)
    if table is None:
        raise HTTPException(404, f"桌号「{table_no}」不存在")
    new_no = body.table_no.strip()
    if new_no != table_no:
        if any(t["table_no"] == new_no for t in tables):
            raise HTTPException(409, f"桌号「{new_no}」已存在")
        # 重命名桌号时同步更新该桌所有宾客
        guests = storage.load_guests()
        for g in guests:
            if g["table_no"] == table_no:
                g["table_no"] = new_no
        storage.save_guests(guests)
    table.update({"table_no": new_no, "label": body.label.strip(), "capacity": body.capacity,
                  "x": body.x, "y": body.y})
    storage.save_tables(tables)
    return table


class SwapTablesIn(BaseModel):
    table_a: str
    table_b: str


@app.post("/api/tables/swap")
def swap_tables(body: SwapTablesIn):
    """交换两桌的备注名、容量和全部宾客，桌号与位置不变。"""
    a_no, b_no = body.table_a.strip(), body.table_b.strip()
    if a_no == b_no:
        raise HTTPException(400, "不能与自己交换")
    tables = storage.load_tables()
    a = next((t for t in tables if t["table_no"] == a_no), None)
    b = next((t for t in tables if t["table_no"] == b_no), None)
    if not a or not b:
        raise HTTPException(404, "桌号不存在")
    a["label"], b["label"] = b["label"], a["label"]
    a["capacity"], b["capacity"] = b["capacity"], a["capacity"]
    storage.save_tables(tables)
    guests = storage.load_guests()
    for g in guests:
        if g["table_no"] == a_no:
            g["table_no"] = b_no
        elif g["table_no"] == b_no:
            g["table_no"] = a_no
    storage.save_guests(guests)
    return {"ok": True}


@app.post("/api/tables/reset-positions")
def reset_table_positions():
    """清空所有桌子坐标 → 平面图恢复按桌号有序的自动排列。"""
    tables = storage.load_tables()
    for t in tables:
        t["x"] = None
        t["y"] = None
    storage.save_tables(tables)
    return {"ok": True, "count": len(tables)}


@app.put("/api/tables/{table_no}/position")
def move_table(table_no: str, body: dict):
    """仅更新桌子坐标（平面图拖动摆位）。"""
    tables = storage.load_tables()
    table = next((t for t in tables if t["table_no"] == table_no), None)
    if table is None:
        raise HTTPException(404, f"桌号「{table_no}」不存在")
    try:
        table["x"] = round(float(body["x"]), 1)
        table["y"] = round(float(body["y"]), 1)
    except (KeyError, TypeError, ValueError):
        raise HTTPException(400, "需要数字坐标 x、y（单位米）")
    storage.save_tables(tables)
    return table


@app.delete("/api/tables/{table_no}")
def delete_table(table_no: str):
    tables = storage.load_tables()
    table = next((t for t in tables if t["table_no"] == table_no), None)
    if table is None:
        raise HTTPException(404, f"桌号「{table_no}」不存在")
    tables.remove(table)
    storage.save_tables(tables)
    # 该桌宾客移入未安排区
    guests = storage.load_guests()
    changed = False
    for g in guests:
        if g["table_no"] == table_no:
            g["table_no"] = ""
            changed = True
    if changed:
        storage.save_guests(guests)
    return {"ok": True}


# ---------- 全局配置 ----------

@app.put("/api/config")
def update_config(body: ConfigIn):
    config = {
        "default_capacity": body.default_capacity,
        "budget_total": body.budget_total,
        "venue_width": body.venue_width,
        "venue_depth": body.venue_depth,
        "table_diameter": body.table_diameter,
        "table_gap": body.table_gap,
        "wedding_date": body.wedding_date,
    }
    storage.save_config(config)
    return config


# 静态页面托管（放最后，避免遮住 /api 路由）
app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
