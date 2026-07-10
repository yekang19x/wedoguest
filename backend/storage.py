"""数据存储层（SQLite）。

所有数据存 data/wedding.db（guests / tables / config 三张表）。
首次启动时：若存在旧版 xlsx/csv 数据文件则自动迁移入库（旧文件保留不动），
否则写入少量示例数据。宾客名单的 Excel 交换通过导入/导出功能完成。
"""

import csv
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DB_PATH = DATA_DIR / "wedding.db"

# 旧版数据文件（仅用于一次性迁移）
LEGACY_GUESTS_XLSX = DATA_DIR / "guests.xlsx"
LEGACY_TABLES_CSV = DATA_DIR / "tables.csv"
LEGACY_CONFIG_CSV = DATA_DIR / "config.csv"
LEGACY_GUEST_HEADER_ALIASES = {"总人数": "预计人数"}

INVITE_STATUSES = ["未发送", "已发送"]
CONFIRM_STATUSES = ["待确认", "已确认", "不参加"]
WECHAT_STATUSES = ["未发送", "已发送"]

DEFAULT_CONFIG = {
    "default_capacity": 10,   # 默认单桌容纳人数
    "budget_total": 100,      # 人数预算
    "venue_width": 18.0,      # 会场宽度（米，横向，舞台所在边）
    "venue_depth": 25.0,      # 会场长度（米，纵向）
    "table_diameter": 1.8,    # 桌子直径（米）
    "table_gap": 0.8,         # 自动排列时相邻桌子的边缘间距（米）
    "wedding_date": "10.25",  # 婚礼日期（显示在标题）
}
CONFIG_INT_KEYS = {"default_capacity", "budget_total"}
CONFIG_STR_KEYS = {"wedding_date"}

_SCHEMA = """
CREATE TABLE IF NOT EXISTS guests (
    id INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    party_size INTEGER NOT NULL DEFAULT 1,
    confirmed_size INTEGER NOT NULL DEFAULT 1,
    family_names TEXT NOT NULL DEFAULT '',
    table_no TEXT NOT NULL DEFAULT '',
    invite_status TEXT NOT NULL DEFAULT '未发送',
    confirm_status TEXT NOT NULL DEFAULT '待确认',
    note TEXT NOT NULL DEFAULT '',
    wechat_sent TEXT NOT NULL DEFAULT '未发送',
    guest_type TEXT NOT NULL DEFAULT '宾客及家属',
    group_id INTEGER
);
CREATE TABLE IF NOT EXISTS tables (
    table_no TEXT PRIMARY KEY,
    label TEXT NOT NULL DEFAULT '',
    capacity INTEGER,
    x REAL,
    y REAL,
    sort_order INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS config (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
"""


def _conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_data_files():
    """建库建表；新库时迁移旧版 xlsx/csv 数据或写入示例数据。"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    fresh = not DB_PATH.exists()
    with _conn() as c:
        c.executescript(_SCHEMA)
    if fresh:
        if LEGACY_GUESTS_XLSX.exists() or LEGACY_TABLES_CSV.exists():
            _migrate_legacy()
        else:
            _seed_demo_data()
    # 版本升级：补新增列
    with _conn() as c:
        cols = {r["name"] for r in c.execute("PRAGMA table_info(guests)").fetchall()}
        if "wechat_sent" not in cols:
            c.execute("ALTER TABLE guests ADD COLUMN wechat_sent TEXT NOT NULL DEFAULT '未发送'")
        if "guest_type" not in cols:
            c.execute("ALTER TABLE guests ADD COLUMN guest_type TEXT NOT NULL DEFAULT '宾客及家属'")
        if "group_id" not in cols:
            c.execute("ALTER TABLE guests ADD COLUMN group_id INTEGER")
    # 配置缺键补默认（版本升级新增配置项时自动补齐）
    config = load_config_raw()
    missing = {k: v for k, v in DEFAULT_CONFIG.items() if k not in config}
    if missing:
        save_config({**missing, **config})


def _seed_demo_data():
    save_tables([
        # 不设坐标：主桌自动排在舞台右侧区域左上角，其余桌按桌号排列
        {"table_no": "主桌", "label": "新人与主宾", "capacity": 12, "x": None, "y": None},
        {"table_no": "1", "label": "男方亲戚", "capacity": None, "x": None, "y": None},
        {"table_no": "2", "label": "女方亲戚", "capacity": None, "x": None, "y": None},
        {"table_no": "3", "label": "同事朋友", "capacity": None, "x": None, "y": None},
    ])
    save_guests([
        {"id": 1, "name": "张伟", "party_size": 3, "confirmed_size": 2, "family_names": "李娜,张小宝",
         "table_no": "1", "invite_status": "已发送", "confirm_status": "已确认", "note": "叔叔一家", "wechat_sent": "已发送"},
        {"id": 2, "name": "王芳", "party_size": 2, "confirmed_size": 2, "family_names": "刘强",
         "table_no": "2", "invite_status": "已发送", "confirm_status": "待确认", "note": "", "wechat_sent": "未发送"},
        {"id": 3, "name": "陈静", "party_size": 1, "confirmed_size": 1, "family_names": "",
         "table_no": "", "invite_status": "未发送", "confirm_status": "待确认", "note": "大学同学", "wechat_sent": "未发送"},
    ])
    save_config(dict(DEFAULT_CONFIG))


# ---------- 宾客 ----------

def load_guests() -> list[dict]:
    with _conn() as c:
        rows = c.execute("SELECT * FROM guests ORDER BY id").fetchall()
    return [dict(r) for r in rows]


def save_guests(guests: list[dict]):
    with _conn() as c:
        c.execute("DELETE FROM guests")
        c.executemany(
            "INSERT INTO guests (id, name, party_size, confirmed_size, family_names,"
            " table_no, invite_status, confirm_status, note, wechat_sent,"
            " guest_type, group_id)"
            " VALUES (:id, :name, :party_size, :confirmed_size, :family_names,"
            " :table_no, :invite_status, :confirm_status, :note, :wechat_sent,"
            " :guest_type, :group_id)",
            [{**g, "wechat_sent": g.get("wechat_sent", "未发送"),
              "guest_type": g.get("guest_type", "宾客及家属"),
              "group_id": g.get("group_id")} for g in guests],
        )


def next_guest_id(guests: list[dict]) -> int:
    return max((g["id"] for g in guests), default=0) + 1


# ---------- 桌子 ----------

def load_tables() -> list[dict]:
    with _conn() as c:
        rows = c.execute(
            "SELECT table_no, label, capacity, x, y FROM tables ORDER BY sort_order, rowid"
        ).fetchall()
    return [dict(r) for r in rows]


def save_tables(tables: list[dict]):
    with _conn() as c:
        c.execute("DELETE FROM tables")
        c.executemany(
            "INSERT INTO tables (table_no, label, capacity, x, y, sort_order)"
            " VALUES (:table_no, :label, :capacity, :x, :y, :sort_order)",
            [{**t, "x": t.get("x"), "y": t.get("y"), "sort_order": i}
             for i, t in enumerate(tables)],
        )


# ---------- 全局配置 ----------

def load_config_raw() -> dict:
    with _conn() as c:
        rows = c.execute("SELECT key, value FROM config").fetchall()
    out = {}
    for r in rows:
        key, val = r["key"], r["value"]
        if key in DEFAULT_CONFIG:
            if key in CONFIG_STR_KEYS:
                out[key] = str(val)
            elif key in CONFIG_INT_KEYS:
                out[key] = int(float(val))
            else:
                out[key] = float(val)
    return out


def load_config() -> dict:
    return {**DEFAULT_CONFIG, **load_config_raw()}


def save_config(config: dict):
    with _conn() as c:
        c.executemany(
            "INSERT INTO config (key, value) VALUES (?, ?)"
            " ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            [(k, str(v)) for k, v in config.items()],
        )


# ---------- 旧版 xlsx/csv 一次性迁移 ----------

def _migrate_legacy():
    tables, config, guests = [], dict(DEFAULT_CONFIG), []
    if LEGACY_TABLES_CSV.exists():
        with open(LEGACY_TABLES_CSV, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                def num(key):
                    raw = (row.get(key) or "").strip()
                    return float(raw) if raw else None
                no = (row.get("桌号") or "").strip()
                if not no:
                    continue
                cap = num("容纳人数")
                tables.append({"table_no": no, "label": (row.get("备注名") or "").strip(),
                               "capacity": int(cap) if cap is not None else None,
                               "x": num("X坐标"), "y": num("Y坐标")})
    if LEGACY_CONFIG_CSV.exists():
        with open(LEGACY_CONFIG_CSV, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                key = (row.get("配置项") or "").strip()
                val = (row.get("值") or "").strip()
                if key in config and val:
                    if key in CONFIG_STR_KEYS:
                        config[key] = val
                    elif key in CONFIG_INT_KEYS:
                        config[key] = int(float(val))
                    else:
                        config[key] = float(val)
    if LEGACY_GUESTS_XLSX.exists():
        ws = load_workbook(LEGACY_GUESTS_XLSX).active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or ()
        col = {}
        for i, h in enumerate(header):
            name = str(h or "").strip()
            col[LEGACY_GUEST_HEADER_ALIASES.get(name, name)] = i

        def cell(vals, name):
            i = col.get(name)
            return vals[i] if i is not None and i < len(vals) else None

        for row in rows:
            if row is None or cell(row, "ID") is None:
                continue
            party_size = int(cell(row, "预计人数") or 1)
            conf_raw = cell(row, "确认人数")
            conf_str = str(conf_raw).strip() if conf_raw is not None else ""
            guests.append({
                "id": int(cell(row, "ID")),
                "name": str(cell(row, "姓名") or "").strip(),
                "party_size": party_size,
                "confirmed_size": int(float(conf_str)) if conf_str else party_size,
                "family_names": str(cell(row, "家属姓名") or "").strip(),
                "table_no": str(cell(row, "桌号") or "").strip(),
                "invite_status": str(cell(row, "邀请函状态") or "未发送").strip(),
                "confirm_status": str(cell(row, "确认状态") or "待确认").strip(),
                "note": str(cell(row, "备注") or "").strip(),
            })
    save_tables(tables)
    save_guests(guests)
    save_config(config)
